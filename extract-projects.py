"""
Extracts project data from DATABASE-DataCenter_Backlash.xlsx
and writes projects-data.js with PROJECTS and PROJECTS_META constants.

The xlsx is read by SHEET POSITION, not by sheet name. Sheets are taken in
consecutive (main, timelines) pairs — currently 4 sheets / 2 pairs:
  1. Philadelphia Fed Region  + PF Timelines
  2. NC VA                    + NC VA Timelines
Any trailing unpaired sheet is ignored with a warning.

Both main sheets share the same Philly-Fed-shaped column layout and are read
through the single FIELD_MAP via header-name lookup (header whitespace is
stripped, so ' Company' and 'Company' both resolve).

Only projects with both a parsable Latitude and Longitude are kept — anything
missing coordinates is dropped from the output (and reported in the summary).

Run: python3 extract-projects.py
"""
import json
import re
import datetime
import urllib.request
from pathlib import Path

import openpyxl
try:
    from shapely.geometry import shape, Point
    HAS_SHAPELY = True
except ImportError:
    HAS_SHAPELY = False

ROOT = Path(__file__).parent
XLSX = ROOT / "DATABASE-DataCenter_Backlash.xlsx"
OUT_JS = ROOT / "projects-data.js"


# Canonical schema — the Philadelphia Fed Region column layout.
# Both sheet pairs are read using this single mapping by header-name lookup,
# so any NC/VA columns that match these names get picked up automatically.
# Anything missing from a sheet falls through to None on that project.
FIELD_MAP = {
    "name": "Project",
    "company": "Company",
    "investmentB": "Investment (B$)",
    "state": "State",
    "county": "County",
    "communities": "Nearby Communities",
    "capacityMw": "Energy Capacity (MW)",
    "acreage": "Acreage",
    "timelineStart": "Timeline - Start",
    "timelineEnd": "Timeline - End",
    "status": "Status",
    "resourceClaims": "Resource Usage Claims",
    "energySources": "Energy Sources",
    "developerPromises": "Developer Promises",
    "concernsCategories": "Comm. Concerns - Categories",
    "articulatedConcerns": "Articulated Concerns",
    "communityPosture": "Comm. Action - Posture",
    "communityIntensity": "Comm. Action - Intensity",
    "communityActionDetails": "Comm. Action - Details",
    "developerAction": "Developer Action",
    "monthRecorded": "Month Recorded",
    "lat": "Latitude",
    "lng": "Longitude",
}

SOURCE_MAP = {
    "projectProposal": "Project Proposal",
    "govtRecords": ["Govt Records"],
    "other": ["Source 3", "Source 4", "Source 5"],
}


# ── drought lookup ───────────────────────────────────────────────────

def _latest_usdm_date():
    """Return the most recent published USDM Tuesday date string (YYYYMMDD)."""
    today = datetime.date.today()
    # USDM releases every Tuesday (weekday 1); find last Tuesday
    days_back = (today.weekday() - 1) % 7
    candidate = today - datetime.timedelta(days=days_back)
    # Walk back up to 4 weeks if today's release isn't up yet
    for _ in range(4):
        url = f"https://droughtmonitor.unl.edu/data/json/usdm_{candidate.strftime('%Y%m%d')}.json"
        try:
            with urllib.request.urlopen(url, timeout=10) as r:
                if r.status == 200:
                    return candidate.strftime('%Y%m%d'), json.loads(r.read())
        except Exception:
            pass
        candidate -= datetime.timedelta(weeks=1)
    return None, None


def build_drought_lookup():
    """Return a function drought_at(lat, lng) → 'D0'..'D4' or None."""
    if not HAS_SHAPELY:
        print("  ⚠ shapely not installed — skipping drought lookup (pip install shapely)")
        return None, None

    print("  Fetching USDM drought polygons…", end=" ", flush=True)
    date_str, geojson = _latest_usdm_date()
    if geojson is None:
        print("failed (no data found)")
        return None, None
    print(f"using {date_str[:4]}-{date_str[4:6]}-{date_str[6:]}")

    # Pre-build shapely geometries sorted by DM level (0→4)
    features = sorted(geojson["features"], key=lambda f: f["properties"]["DM"])
    polygons = [(f["properties"]["DM"], shape(f["geometry"])) for f in features]
    DM_LABELS = {0: "D0", 1: "D1", 2: "D2", 3: "D3", 4: "D4"}

    def drought_at(lat, lng):
        pt = Point(lng, lat)
        result = None
        for dm, geom in polygons:
            if geom.contains(pt):
                result = dm  # higher DM wins; list is sorted ascending
        return DM_LABELS.get(result) if result is not None else "None"

    return drought_at, date_str


# ── helpers ──────────────────────────────────────────────────────────

def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def slugify(name):
    s = re.sub(r"[^\w\s-]", "", name.lower())
    s = re.sub(r"\s+", "-", s).strip("-")
    return s


def normalize_intensity(v):
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ('none', ''):
        return 'None'
    if s in ('low',):
        return 'Low'
    if s in ('mod', 'moderate', 'medium', 'med'):
        return 'Moderate'
    if s in ('high',):
        return 'High'
    return str(v).strip()


def parse_date(v):
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    if isinstance(v, (int, float)):
        # NC VA "Timeline" column often holds a bare year like 2028
        if 1900 <= v <= 2100 and v == int(v):
            return str(int(v))
        return None
    if isinstance(v, str):
        return v.strip() or None
    return None


def split_sources(v):
    if not v:
        return []
    return [u.strip() for u in str(v).split(";") if u.strip()]


def build_header_index(headers):
    """Map cleaned header string → column index. Skips None / blank headers."""
    out = {}
    for i, h in enumerate(headers):
        h_clean = clean(h)
        if h_clean and h_clean not in out:
            out[h_clean] = i
    return out


def resolve_col(spec, header_index):
    """Translate a column spec into a column index (or None)."""
    if spec is None:
        return None
    if isinstance(spec, int):
        return spec
    return header_index.get(spec)


def cell(row, spec, header_index):
    idx = resolve_col(spec, header_index)
    if idx is None or idx >= len(row):
        return None
    return clean(row[idx])


# ── extraction ───────────────────────────────────────────────────────

def extract_pair(main_ws, timelines_ws):
    rows = list(main_ws.iter_rows(values_only=True))
    if not rows:
        return []
    header_index = build_header_index(rows[0])
    field_map = FIELD_MAP
    source_map = SOURCE_MAP

    projects = []
    for r in rows[1:]:
        # A fully-blank row marks the end of the table — stop reading here.
        # (Anything below is treated as a deliberate separator/scratch area,
        # even if a later row happens to have content.)
        if all(clean(v) is None for v in r):
            break

        name = cell(r, field_map["name"], header_index)
        if not name:
            continue
        clean_name = name.lstrip("✶* ").strip()

        proj = {
            "id": slugify(clean_name),
            "name": clean_name,
        }
        for target, spec in field_map.items():
            if target == "name":
                continue
            val = cell(r, spec, header_index)
            if target in ("timelineStart", "timelineEnd", "monthRecorded"):
                val = parse_date(val) if val is not None else None
            elif target == "communityIntensity":
                val = normalize_intensity(val)
            elif target in ("lat", "lng"):
                if val is None:
                    pass
                else:
                    try:
                        val = float(val)
                    except (TypeError, ValueError):
                        val = None  # unparsable coords get dropped silently
            proj[target] = val

        proj_proposal = cell(r, source_map.get("projectProposal"), header_index)
        govt_records = []
        for spec in source_map.get("govtRecords") or []:
            govt_records.extend(split_sources(cell(r, spec, header_index)))
        other_sources = []
        for spec in source_map.get("other") or []:
            other_sources.extend(split_sources(cell(r, spec, header_index)))
        proj["sources"] = {
            "projectProposal": proj_proposal,
            "govtRecords": govt_records,
            "other": other_sources,
        }

        # Stub detection: count fields with real content
        substantive = sum(
            1 for k, v in proj.items()
            if k not in ("id", "name", "featured", "sources") and v not in (None, "", [])
        )
        proj["stub"] = substantive <= 1

        # placeholders for fields not in the spreadsheet yet
        proj["coordsPrecision"] = None
        proj["timeline"] = []

        projects.append(proj)

    attach_timelines(timelines_ws, projects)
    return projects


def attach_timelines(ws, projects):
    if ws is None:
        return
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return

    # Map cleaned project name → column in the timeline sheet
    col_by_name = {}
    for j, v in enumerate(rows[0]):
        n = clean(v)
        if n:
            cleaned = n.lstrip("✶* ").strip()
            col_by_name.setdefault(cleaned, j)

    # Walk rows looking for (label, date, source) triples
    triples = []
    i = 1
    while i < len(rows):
        label = clean(rows[i][0])
        if label and ("Proposal/Announcement" in str(label) or "Development" in str(label)):
            triples.append((
                i,
                i + 1 if i + 1 < len(rows) else None,
                i + 2 if i + 2 < len(rows) else None,
                "Proposal/Announcement" in str(label),
            ))
            i += 3
        else:
            i += 1

    by_name = {p["name"]: p for p in projects}
    for name, col in col_by_name.items():
        proj = by_name.get(name)
        if not proj:
            continue
        for (lr, dr, sr, is_proposal) in triples:
            label = clean(rows[lr][col]) if lr is not None else None
            date = parse_date(rows[dr][col]) if dr is not None else None
            source = clean(rows[sr][col]) if sr is not None else None
            if not (label or date or source):
                continue
            proj["timeline"].append({
                "date": date,
                "label": label,
                "isProposal": is_proposal,
                "source": source,
            })
        proj["timeline"].sort(key=lambda e: (e["date"] is None, e["date"] or ""))


# ── main ─────────────────────────────────────────────────────────────

def main():
    drought_at, drought_date = build_drought_lookup()

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    sheets = wb.worksheets
    if len(sheets) < 2:
        raise SystemExit("Workbook needs at least one (main, timelines) sheet pair.")

    # Walk consecutive (main, timelines) pairs: (0,1), (2,3), ...
    pairs = [(sheets[i], sheets[i + 1]) for i in range(0, len(sheets) - 1, 2)]
    if len(sheets) % 2 != 0:
        print(f"  ⚠ Ignoring trailing unpaired sheet: '{sheets[-1].title}'")

    all_projects = []
    pair_stats = []
    for main_ws, timelines_ws in pairs:
        projects = extract_pair(main_ws, timelines_ws)
        all_projects.extend(projects)
        pair_stats.append({
            "mainSheet": main_ws.title,
            "timelinesSheet": timelines_ws.title,
            "projectCount": len(projects),
            "fullCount": sum(1 for p in projects if not p["stub"]),
        })

    # Keep only projects with parsable coordinates on both axes
    before = len(all_projects)
    all_projects = [p for p in all_projects if p.get("lat") is not None and p.get("lng") is not None]
    dropped = before - len(all_projects)

    # Inject drought level right after coords
    for p in all_projects:
        dl = drought_at(p["lat"], p["lng"]) if drought_at else None
        # Insert droughtLevel immediately after lng in the dict
        items = list(p.items())
        lng_idx = next((i for i, (k, _) in enumerate(items) if k == "lng"), None)
        if lng_idx is not None:
            items.insert(lng_idx + 1, ("droughtLevel", dl))
            p.clear()
            p.update(items)
        else:
            p["droughtLevel"] = dl
    if dropped:
        print(f"  ⚠ Dropped {dropped} project(s) without coordinates")

    # Detect id collisions across pairs
    seen = {}
    for p in all_projects:
        if p["id"] in seen:
            print(f"  ⚠ duplicate id '{p['id']}' — '{p['name']}' clashes with '{seen[p['id']]}'")
        else:
            seen[p["id"]] = p["name"]

    payload_meta = {
        "generatedAt": datetime.datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "sourceFile": XLSX.name,
        "sheetsRead": pair_stats,
        "droppedNoCoords": dropped,
        "droughtDate": drought_date,
    }

    js = (
        "// Auto-generated from DATABASE-DataCenter_Backlash.xlsx\n"
        "// Run extract-projects.py to regenerate.\n"
        "// Do not edit by hand.\n\n"
        f"const PROJECTS_META = {json.dumps(payload_meta, indent=2, ensure_ascii=False)};\n\n"
        f"const PROJECTS = {json.dumps(all_projects, indent=2, ensure_ascii=False)};\n"
    )
    OUT_JS.write_text(js, encoding="utf-8")

    total_events = sum(len(p["timeline"]) for p in all_projects)
    total_full = sum(s["fullCount"] for s in pair_stats)
    print(f"Wrote {OUT_JS.name}")
    for s in pair_stats:
        stubs = s["projectCount"] - s["fullCount"]
        print(f"  {s['mainSheet']!r} + {s['timelinesSheet']!r}: {s['projectCount']} projects ({s['fullCount']} full, {stubs} stub)")
    print(f"  {len(all_projects)} projects with coordinates kept ({total_full} full), {total_events} timeline events")


if __name__ == "__main__":
    main()
